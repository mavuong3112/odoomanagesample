from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64
import io
import re
from openpyxl import load_workbook


class ImportQuestionWizard(models.TransientModel):
    _name = "poll.import.question.wizard"
    _description = "Import Poll Questions from Excel"

    file = fields.Binary(string="Tệp Excel", required=True, help="Tệp .xlsx tải xuống từ Google Form")
    filename = fields.Char(string="Tên tệp")

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        res.setdefault("filename", "google_form_responses.xlsx")
        return res

    def _get_option_indices(self, headers):
        pat = re.compile(r'^Option\d+$', re.IGNORECASE)
        return [i for i, h in enumerate(headers) if pat.match(str(h or '').strip())]

    def action_import(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("Vui lòng chọn tệp Excel."))

        wb_buffer = io.BytesIO(base64.b64decode(self.file))
        try:
            wb = load_workbook(wb_buffer, data_only=True)
        except Exception as e:
            raise UserError(_("Không thể đọc file Excel: %s") % str(e))

        sheet = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(sheet.iter_rows(min_row=1, max_row=1))]

        if 'Question' not in headers:
            raise UserError(_('File phải có cột "Question".'))

        q_idx = headers.index('Question')
        opt_indices = self._get_option_indices(headers)
        if not opt_indices:
            raise UserError(_('Không tìm thấy cột "Option1", "Option2", …'))

        corr_idx = headers.index('CorrectOption') if 'CorrectOption' in headers else None

        PollQuestion = self.env["poll.question"]
        PollAnswer = self.env["poll.answer"]
        created = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = list(row)
            question_text = values[q_idx] if q_idx < len(values) else ""
            if not question_text:
                continue

            correct_val = values[corr_idx] if corr_idx is not None and corr_idx < len(values) else None

            question = PollQuestion.create({
                "name": question_text,
                "description": question_text,
            })

            for idx in opt_indices:
                if idx >= len(values):
                    continue
                opt_text = values[idx]
                if not opt_text:
                    continue
                ans = PollAnswer.create({
                    "name": opt_text,
                    "question_id": question.id,
                })
                if correct_val and str(correct_val).strip().lower() == headers[idx].lower():
                    ans.count = 1

            created += 1

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Import thành công"),
                "message": _("Đã tạo %s câu hỏi mới.") % created,
                "sticky": False,
                "type": "success",
            },
        }